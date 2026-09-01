"""
Sprint 3 — Day 17: Generates output/screener_output.xlsx — one sheet per
preset, 20 KPI columns, sorted by composite score descending, with
green/red conditional formatting against each preset's own thresholds.
"""
import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))
from screener.engine import load_config, load_screener_universe, run_all_presets
from screener.composite_score import compute_composite_scores

ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "db" / "nifty100.db"
OUTPUT = ROOT / "output"

DISPLAY_COLS = [
    "company_id", "company_name", "broad_sector",
    "return_on_equity_pct", "roce_pct", "net_profit_margin_pct",
    "debt_to_equity", "interest_coverage", "icr_label",
    "asset_turnover", "free_cash_flow_cr", "cfo_quality_label",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "pe_ratio", "pb_ratio", "dividend_yield_pct",
    "dividend_payout_ratio_pct", "sales", "composite_quality_score",
]

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True)


def threshold_pass(value, metric, condition):
    if pd.isna(value):
        return None
    if "min" in condition:
        return value >= condition["min"]
    if "max" in condition:
        return value <= condition["max"]
    if "equals" in condition:
        return value == condition["equals"]
    return None


def write_sheet(writer, sheet_name, df, filters):
    cols = [c for c in DISPLAY_COLS if c in df.columns]
    export_df = df[cols].copy()
    export_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    for c_idx, col in enumerate(cols, start=1):
        ws.cell(row=1, column=c_idx).font = HEADER_FONT
        if col not in filters:
            continue
        for r_idx, value in enumerate(export_df[col], start=2):
            passed = threshold_pass(value, col, filters[col])
            if passed is True:
                ws.cell(row=r_idx, column=c_idx).fill = GREEN
            elif passed is False:
                ws.cell(row=r_idx, column=c_idx).fill = RED

    for c_idx, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = max(14, len(col) + 2)


def main():
    conn = sqlite3.connect(DB_PATH)
    config = load_config()
    df = load_screener_universe(conn)
    df = compute_composite_scores(df)
    results = run_all_presets(df, config, conn=conn)

    with pd.ExcelWriter(OUTPUT / "screener_output.xlsx", engine="openpyxl") as writer:
        for key, res_df in results.items():
            label = config["presets"][key]["label"]
            filters = config["presets"][key]["filters"]
            write_sheet(writer, label[:31], res_df, filters)  # Excel sheet name limit = 31 chars
            print(f"{label:25} -> {len(res_df):3} companies written")

    conn.close()


if __name__ == "__main__":
    main()
